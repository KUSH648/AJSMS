import os
import openpyxl
from datetime import datetime
from flask import current_app

def get_export_path(filename):
    export_dir = os.path.join('static', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    return os.path.join(export_dir, filename)

def style_excel_sheet(ws, headers):
    """Apply basic styling to header row."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='D4AF37', end_color='B8860B', fill_type='solid')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 25
    for col in ws.columns:
        for c in col[1:]:
            c.alignment = Alignment(horizontal='center')

def export_inventory_to_excel(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = ["Name", "Category", "Material", "Purity", "Weight (gm)", "Selling Price", "Stock Qty", "Barcode"]
    ws.append(headers)
    for item in items:
        ws.append([
            item.get('name', ''),
            item.get('category', ''),
            item.get('material', ''),
            item.get('purity', ''),
            item.get('weight_gm', 0.0),
            item.get('selling_price', 0.0),
            item.get('stock_qty', 0),
            item.get('barcode', '')
        ])
    style_excel_sheet(ws, headers)
    filename = f"inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = get_export_path(filename)
    wb.save(filepath)
    return os.path.abspath(filepath)

def export_bills_to_excel(bills):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills"
    headers = ["Bill No", "Date", "Customer", "Subtotal", "GST", "Discount", "Total", "Status"]
    ws.append(headers)
    for b in bills:
        ws.append([
            b.get('bill_number', ''),
            b.get('bill_date', ''),
            b.get('customer_name', ''),
            b.get('subtotal', 0.0),
            b.get('gst_amount', 0.0),
            b.get('discount', 0.0),
            b.get('total_amount', 0.0),
            b.get('status', '')
        ])
    style_excel_sheet(ws, headers)
    filename = f"bills_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = get_export_path(filename)
    wb.save(filepath)
    return os.path.abspath(filepath)

def export_customers_to_excel(customers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["Name", "Phone", "Email", "Loyalty Points", "Created At"]
    ws.append(headers)
    for c in customers:
        ws.append([
            c.get('name', ''),
            c.get('phone', ''),
            c.get('email', ''),
            c.get('loyalty_points', 0),
            c.get('created_at', '')
        ])
    style_excel_sheet(ws, headers)
    filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = get_export_path(filename)
    wb.save(filepath)
    return os.path.abspath(filepath)

def export_sales_to_excel(sales, date_from='', date_to=''):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    headers = ["Date", "Customer", "Employee", "Subtotal", "GST", "Total", "Payment Method"]
    ws.append(headers)
    for s in sales:
        ws.append([
            s.get('sale_date', ''),
            s.get('customer_name', ''),
            s.get('employee_name', ''),
            s.get('subtotal', 0.0),
            s.get('gst_amount', 0.0),
            s.get('total_amount', 0.0),
            s.get('payment_method', '')
        ])
    style_excel_sheet(ws, headers)
    filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = get_export_path(filename)
    wb.save(filepath)
    return os.path.abspath(filepath)

def export_attendance_to_excel(data, date_from='', date_to=''):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Employee", "Date", "Check In", "Check Out", "Status"]
    ws.append(headers)
    for a in data:
        ws.append([
            a.get('employee_name', a.get('name', '')),
            a.get('work_date', a.get('date', '')),
            a.get('check_in', '09:00'),
            a.get('check_out', '18:00'),
            a.get('status', '')
        ])
    style_excel_sheet(ws, headers)
    filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = get_export_path(filename)
    wb.save(filepath)
    return os.path.abspath(filepath)
